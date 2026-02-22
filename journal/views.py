from datetime import timedelta
from decimal import Decimal
from datetime import date as date_type

from django.conf import settings
from django.utils import timezone
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.urls import reverse_lazy
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator

from django.contrib.contenttypes.models import ContentType

from .models import ActivityLog, JournalEntry, JournalEntryLine, ApprovalRequest, UserRequest
from .forms import JournalEntryForm, JournalEntryLineFormSet, SalesForm, ExpenseForm, UserRequestForm, UserResetPasswordForm
from .mixins import RoleRequiredMixin, user_in_role
from .utils import log_activity


def _user_can_bypass_approval(user):
    """Admin or Manager can bypass 12-hour approval."""
    from .mixins import user_in_role
    return user_in_role(user, 'admin') or user_in_role(user, 'manager')


class HomeView(RoleRequiredMixin, TemplateView):
    template_name = 'journal/home.html'
    allowed_roles = ['admin', 'manager', 'staff', 'viewer']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['show_approval_queue'] = _user_can_bypass_approval(self.request.user)
        context['show_user_approval'] = user_in_role(self.request.user, 'admin')
        return context


def _apply_entry_period_filter(queryset, period, date_str, today):
    """Apply date-period filter to a JournalEntry queryset. Returns filtered qs."""
    if not period:
        return queryset
    if period == 'day' and date_str:
        try:
            parsed = date_type.fromisoformat(date_str)
            return queryset.filter(date=parsed)
        except (ValueError, TypeError):
            return queryset
    if period == 'week':
        start = today - timedelta(days=6)
        return queryset.filter(date__gte=start, date__lte=today)
    if period == 'month':
        return queryset.filter(date__year=today.year, date__month=today.month)
    if period == 'year':
        return queryset.filter(date__year=today.year)
    return queryset


def _apply_date_range_filter(queryset, date_from_str, date_to_str):
    """Apply date range filter. Returns filtered qs."""
    if date_from_str:
        try:
            parsed = date_type.fromisoformat(date_from_str.strip())
            queryset = queryset.filter(date__gte=parsed)
        except (ValueError, TypeError):
            pass
    if date_to_str:
        try:
            parsed = date_type.fromisoformat(date_to_str.strip())
            queryset = queryset.filter(date__lte=parsed)
        except (ValueError, TypeError):
            pass
    return queryset


class EntryListView(RoleRequiredMixin, ListView):
    model = JournalEntry
    context_object_name = 'entries'
    template_name = 'journal/entry_list.html'
    paginate_by = 20
    allowed_roles = ['admin', 'manager', 'staff', 'viewer']

    def get_queryset(self):
        qs = JournalEntry.objects.active().order_by('-date', '-created_at')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        return _apply_date_range_filter(qs, date_from, date_to)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        entries = context.get('entries') or context.get('object_list') or []
        entry_ids = [e.id for e in entries]
        if entry_ids:
            ct = ContentType.objects.get_for_model(JournalEntry)
            logs = ActivityLog.objects.filter(
                content_type=ct, object_id__in=entry_ids
            ).select_related('user').order_by('-timestamp')
            entry_logs = {}
            for log in logs:
                entry_logs.setdefault(log.object_id, []).append(log)
            for entry in entries:
                entry.activity_logs = entry_logs.get(entry.id, [])
        else:
            for entry in entries:
                entry.activity_logs = []
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


class RecordsExportView(RoleRequiredMixin, View):
    """Export filtered records to Excel."""
    allowed_roles = ['admin', 'manager', 'staff', 'viewer']

    def get(self, request):
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ModuleNotFoundError as e:
            messages.error(
                request,
                'Excel export requires openpyxl. Run: pip install openpyxl'
            )
            return redirect('journal:entry_list')

        qs = JournalEntry.objects.active().order_by('-date', '-created_at')
        today = timezone.now().date()
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        qs = _apply_date_range_filter(qs, date_from, date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Records'

        thin_border = Side(style='thin', color='CCCCCC')
        header_border = Side(style='thin', color='999999')
        header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        # Header row: Pure Herb — Records (branded, centered, padded)
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 32
        brand_cell = ws['A1']
        brand_cell.value = 'Pure Herb — Records'
        brand_cell.font = Font(bold=True, size=14)
        brand_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Padding row
        ws.row_dimensions[2].height = 12

        # Column headers (same order as website table)
        headers = ['Date', 'Reference', 'Type', 'Description', 'Total (AED)', 'Created by']
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='right' if col == 5 else 'left', vertical='center')
            cell.border = Border(top=header_border, bottom=header_border, left=header_border, right=header_border)

        # Column widths for readability
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14

        for row_idx, entry in enumerate(qs, header_row + 1):
            date_str = f"{entry.date.strftime('%b')} {entry.date.day}, {entry.date.strftime('%Y')}" if hasattr(entry.date, 'strftime') else str(entry.date)
            ws.cell(row=row_idx, column=1, value=date_str)
            ws.cell(row=row_idx, column=2, value=entry.reference or '—')
            ws.cell(row=row_idx, column=3, value=entry.get_entry_type_display())
            ws.cell(row=row_idx, column=4, value=entry.description or '')
            amt_cell = ws.cell(row=row_idx, column=5, value=float(entry.total_debit))
            amt_cell.number_format = '#,##0.00'
            amt_cell.alignment = Alignment(horizontal='right')
            ws.cell(row=row_idx, column=6, value=entry.created_by.username if entry.created_by else '—')
            for c in range(1, 7):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = Border(bottom=thin_border, left=thin_border, right=thin_border)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'records-{today.isoformat()}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class EntryDetailView(RoleRequiredMixin, DetailView):
    model = JournalEntry
    context_object_name = 'entry'
    template_name = 'journal/entry_detail.html'
    allowed_roles = ['admin', 'manager', 'staff', 'viewer']


class EntryCreateView(RoleRequiredMixin, CreateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'journal/entry_form.html'
    success_url = reverse_lazy('journal:home')
    allowed_roles = ['admin', 'manager', 'staff']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'line_formset' not in kwargs:
            if self.request.POST:
                context['line_formset'] = JournalEntryLineFormSet(
                    self.request.POST,
                    instance=self.object if self.object else None
                )
            else:
                context['line_formset'] = JournalEntryLineFormSet(
                    instance=self.object if self.object else None
                )
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        line_formset = JournalEntryLineFormSet(
            self.request.POST,
            instance=self.object
        )
        if line_formset.is_valid():
            line_formset.save()
            log_activity(self.request.user, 'created', self.object)
            messages.success(self.request, 'Record created successfully.')
            return redirect(self.success_url)
        else:
            self.object.delete()
            return self.render_to_response(
                self.get_context_data(form=form, line_formset=line_formset)
            )


class EntryUpdateView(RoleRequiredMixin, UpdateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'journal/entry_form.html'
    context_object_name = 'entry'
    success_url = reverse_lazy('journal:home')
    allowed_roles = ['admin', 'manager', 'staff']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['line_formset'] = JournalEntryLineFormSet(
                self.request.POST,
                instance=self.object
            )
        else:
            context['line_formset'] = JournalEntryLineFormSet(
                instance=self.object
            )
        return context

    def form_valid(self, form):
        # 12-hour approval: Staff needs approval for entries older than 12h
        cutoff = timezone.now() - timedelta(hours=12)
        created_at = self.object.created_at
        if timezone.is_naive(created_at):
            created_at = timezone.make_aware(created_at, timezone.get_default_timezone())
        if created_at < cutoff and not _user_can_bypass_approval(self.request.user):
            # Create approval request and redirect (only if formset is valid)
            line_formset = JournalEntryLineFormSet(self.request.POST, instance=self.object)
            if not line_formset.is_valid():
                return self.form_invalid(form)
            payload = {
                'date': str(form.cleaned_data.get('date')),
                'reference': form.cleaned_data.get('reference', ''),
                'description': form.cleaned_data.get('description', ''),
            }
            lines = []
            for f in line_formset.forms:
                if f.cleaned_data and not f.cleaned_data.get('DELETE'):
                    lines.append({
                        'account': f.cleaned_data.get('account', ''),
                        'debit': str(f.cleaned_data.get('debit') or 0),
                        'credit': str(f.cleaned_data.get('credit') or 0),
                        'memo': f.cleaned_data.get('memo', ''),
                    })
            payload['lines'] = lines
            ApprovalRequest.objects.create(
                entry=self.object,
                action='update',
                requested_by=self.request.user,
                status='pending',
                payload=payload
            )
            messages.info(self.request, 'Your update request has been submitted for approval.')
            return redirect('journal:approval_pending')
        line_formset = JournalEntryLineFormSet(
            self.request.POST,
            instance=self.object
        )
        if line_formset.is_valid():
            self.object.updated_by = self.request.user
            form.save()
            line_formset.save()
            log_activity(self.request.user, 'updated', self.object)
            messages.success(self.request, 'Record updated successfully.')
            return redirect(self.success_url)
        else:
            return self.form_invalid(form)

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))


class EntryDeleteView(RoleRequiredMixin, DeleteView):
    model = JournalEntry
    context_object_name = 'entry'
    template_name = 'journal/entry_confirm_delete.html'
    success_url = reverse_lazy('journal:home')
    allowed_roles = ['admin', 'manager', 'staff']

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        # 12-hour approval: Staff needs approval for entries older than 12h
        cutoff = timezone.now() - timedelta(hours=12)
        created_at = self.object.created_at
        if timezone.is_naive(created_at):
            created_at = timezone.make_aware(created_at, timezone.get_default_timezone())
        if created_at < cutoff and not _user_can_bypass_approval(request.user):
            ApprovalRequest.objects.create(
                entry=self.object,
                action='delete',
                requested_by=request.user,
                status='pending',
                payload={'description': self.object.description, 'date': str(self.object.date)}
            )
            messages.info(request, 'Your delete request has been submitted for approval.')
            return redirect('journal:approval_pending')
        # Soft delete
        from django.utils import timezone as tz
        self.object.deleted_at = tz.now()
        self.object.updated_by = request.user
        self.object.save()
        log_activity(request.user, 'deleted', self.object, extra={'description': self.object.description})
        messages.success(request, 'Record deleted.')
        return redirect(self.success_url)


def _create_sale_from_form(form, user):
    cd = form.cleaned_data
    entry = JournalEntry.objects.create(
        date=cd['date'],
        reference=cd.get('reference', ''),
        description=cd['description'],
        entry_type='sale',
        created_by=user
    )
    amount = cd['amount']
    payment_method = cd.get('payment_method', 'cash')
    debit_account = 'Cash' if payment_method == 'cash' else 'Card'
    JournalEntryLine.objects.create(
        entry=entry, account=debit_account, debit=amount, credit=Decimal('0')
    )
    JournalEntryLine.objects.create(
        entry=entry, account='Revenue', debit=Decimal('0'), credit=amount
    )
    log_activity(user, 'created', entry)
    return entry


def _create_expense_from_form(form, user):
    cd = form.cleaned_data
    entry = JournalEntry.objects.create(
        date=cd['date'],
        reference=cd.get('reference', ''),
        description=cd['description'],
        entry_type='expense',
        created_by=user
    )
    amount = cd['amount']
    category = cd['category']
    payment_method = cd.get('payment_method', 'cash')
    credit_account = 'Cash' if payment_method == 'cash' else 'Card'
    JournalEntryLine.objects.create(
        entry=entry, account=category, debit=amount, credit=Decimal('0')
    )
    JournalEntryLine.objects.create(
        entry=entry, account=credit_account, debit=Decimal('0'), credit=amount
    )
    log_activity(user, 'created', entry)
    return entry


class SaleFormView(RoleRequiredMixin, View):
    """GET: show Record Sale form. POST: create sale and redirect."""
    allowed_roles = ['admin', 'manager', 'staff']

    def get(self, request):
        from datetime import date
        form = SalesForm(initial={'date': date.today()})
        return render(request, 'journal/entry_sales_form.html', {'form': form})

    def post(self, request):
        form = SalesForm(request.POST)
        if form.is_valid():
            _create_sale_from_form(form, request.user)
            messages.success(request, 'Sale recorded successfully.')
            return redirect('journal:home')
        return render(request, 'journal/entry_sales_form.html', {'form': form})


class ExpenseFormView(RoleRequiredMixin, View):
    """GET: show Record Expense form. POST: create expense and redirect."""
    allowed_roles = ['admin', 'manager', 'staff']

    def get(self, request):
        from datetime import date
        form = ExpenseForm(initial={'date': date.today()})
        return render(request, 'journal/entry_expense_form.html', {'form': form})

    def post(self, request):
        form = ExpenseForm(request.POST)
        if form.is_valid():
            _create_expense_from_form(form, request.user)
            messages.success(request, 'Expense recorded successfully.')
            return redirect('journal:home')
        return render(request, 'journal/entry_expense_form.html', {'form': form})


class ApprovalPendingView(RoleRequiredMixin, TemplateView):
    template_name = 'journal/approval_pending.html'
    allowed_roles = ['admin', 'manager', 'staff']


class ApprovalQueueView(RoleRequiredMixin, ListView):
    model = ApprovalRequest
    context_object_name = 'requests'
    template_name = 'journal/approval_queue.html'
    allowed_roles = ['admin', 'manager']

    def get_queryset(self):
        return ApprovalRequest.objects.filter(status='pending')


class ApprovalRequestApproveView(RoleRequiredMixin, View):
    allowed_roles = ['admin', 'manager']

    def post(self, request, pk):
        ar = get_object_or_404(ApprovalRequest, pk=pk, status='pending')
        action = request.POST.get('action')  # 'approve' or 'reject'
        if action == 'approve':
            entry = ar.entry
            if ar.action == 'delete':
                entry.deleted_at = timezone.now()
                entry.updated_by = request.user
                entry.save()
                log_activity(request.user, 'deleted', entry, extra={'approved': True})
            else:  # update
                payload = ar.payload or {}
                from datetime import datetime
                if 'date' in payload:
                    entry.date = datetime.strptime(payload['date'], '%Y-%m-%d').date()
                entry.reference = payload.get('reference', '')
                entry.description = payload.get('description', '')
                entry.updated_by = request.user
                entry.save()
                if 'lines' in payload:
                    entry.lines.all().delete()
                    for line in payload['lines']:
                        JournalEntryLine.objects.create(
                            entry=entry,
                            account=line.get('account', ''),
                            debit=Decimal(line.get('debit', 0)),
                            credit=Decimal(line.get('credit', 0)),
                            memo=line.get('memo', '')
                        )
                log_activity(request.user, 'updated', entry, extra={'approved': True})
            ar.status = 'approved'
            ar.approved_by = request.user
            ar.approved_at = timezone.now()
            ar.save()
            messages.success(request, 'Request approved.')
        else:
            ar.status = 'rejected'
            ar.approved_by = request.user
            ar.approved_at = timezone.now()
            ar.save()
            messages.info(request, 'Request rejected.')
        return redirect('journal:approval_queue')


class UserRequestCreateView(RoleRequiredMixin, CreateView):
    model = UserRequest
    form_class = UserRequestForm
    template_name = 'journal/user_request_form.html'
    success_url = reverse_lazy('journal:user_request_pending')
    allowed_roles = ['admin', 'manager']

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if user_in_role(self.request.user, 'admin'):
            from django.contrib.auth.models import User, Group
            new_user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data.get('email') or '',
                password=form.cleaned_data.get('password') or User.objects.make_random_password()
            )
            role_to_group = {'admin': 'Admin', 'manager': 'Manager', 'staff': 'Staff', 'viewer': 'Viewer'}
            group_name = role_to_group.get(form.cleaned_data['role'], form.cleaned_data['role'].capitalize())
            try:
                group = Group.objects.get(name=group_name)
            except Group.DoesNotExist:
                group = None
            if group:
                new_user.groups.add(group)
            UserRequest.objects.create(
                requested_by=self.request.user,
                username=form.cleaned_data['username'],
                email=form.cleaned_data.get('email') or '',
                role=form.cleaned_data['role'],
                password='',
                status='approved',
                approved_by=self.request.user,
                approved_at=timezone.now()
            )
            messages.success(self.request, f'User {form.cleaned_data["username"]} created.')
            return redirect('journal:user_request_pending')
        form.instance.requested_by = self.request.user
        form.instance.status = 'pending'
        return super().form_valid(form)


class UserRequestPendingView(RoleRequiredMixin, TemplateView):
    template_name = 'journal/user_request_pending.html'
    allowed_roles = ['admin', 'manager']


class UserRequestApprovalView(RoleRequiredMixin, ListView):
    model = UserRequest
    context_object_name = 'requests'
    template_name = 'journal/user_request_approval.html'
    allowed_roles = ['admin']

    def get_queryset(self):
        return UserRequest.objects.filter(status='pending')


class UserRequestApproveView(RoleRequiredMixin, View):
    allowed_roles = ['admin']

    def post(self, request, pk):
        ur = get_object_or_404(UserRequest, pk=pk, status='pending')
        action = request.POST.get('action')
        if action == 'approve':
            from django.contrib.auth.models import User, Group
            user = User.objects.create_user(
                username=ur.username,
                email=ur.email or '',
                password=ur.password or User.objects.make_random_password()
            )
            role_to_group = {'admin': 'Admin', 'manager': 'Manager', 'staff': 'Staff', 'viewer': 'Viewer'}
            group_name = role_to_group.get(ur.role, ur.role.capitalize())
            try:
                group = Group.objects.get(name=group_name)
            except Group.DoesNotExist:
                group = None
            if group:
                user.groups.add(group)
            ur.status = 'approved'
            ur.approved_by = request.user
            ur.approved_at = timezone.now()
            ur.save()
            messages.success(request, f'User {ur.username} created.')
        else:
            ur.status = 'rejected'
            ur.approved_by = request.user
            ur.approved_at = timezone.now()
            ur.save()
            messages.info(request, 'User request rejected.')
        return redirect('journal:user_request_approval')


class UserManagementListView(RoleRequiredMixin, ListView):
    """Admin-only list of all users."""
    allowed_roles = ['admin']
    template_name = 'journal/user_management.html'
    context_object_name = 'users'
    paginate_by = 20

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        qs = User.objects.all().order_by('username').prefetch_related('groups')
        hardwired = getattr(settings, 'HARDWIRED_USERNAME', None)
        if hardwired:
            qs = qs.exclude(username=hardwired)
        return qs


class UserResetPasswordView(RoleRequiredMixin, View):
    """Admin-only view to reset a user's password."""
    allowed_roles = ['admin']

    def get(self, request, pk):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        target = get_object_or_404(User, pk=pk)
        if getattr(settings, 'HARDWIRED_USERNAME', None) and target.username == settings.HARDWIRED_USERNAME:
            messages.error(request, 'This user cannot be modified.')
            return redirect('journal:user_management')
        form = UserResetPasswordForm()
        return render(request, 'journal/user_reset_password.html', {'target_user': target, 'form': form})

    def post(self, request, pk):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        target = get_object_or_404(User, pk=pk)
        if getattr(settings, 'HARDWIRED_USERNAME', None) and target.username == settings.HARDWIRED_USERNAME:
            messages.error(request, 'This user cannot be modified.')
            return redirect('journal:user_management')
        form = UserResetPasswordForm(request.POST)
        if form.is_valid():
            target.set_password(form.cleaned_data['password1'])
            target.save()
            messages.success(request, f'Password for {target.username} has been reset.')
            return redirect('journal:user_management')
        return render(request, 'journal/user_reset_password.html', {'target_user': target, 'form': form})


class UserToggleLockView(RoleRequiredMixin, View):
    """Admin-only view to lock/unlock a user (toggle is_active)."""
    allowed_roles = ['admin']

    def post(self, request, pk):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        target = get_object_or_404(User, pk=pk)
        if getattr(settings, 'HARDWIRED_USERNAME', None) and target.username == settings.HARDWIRED_USERNAME:
            messages.error(request, 'This user cannot be locked or modified.')
            return redirect('journal:user_management')
        if target.pk == request.user.pk:
            messages.error(request, 'You cannot lock yourself.')
            return redirect('journal:user_management')
        admin_count = User.objects.filter(groups__name='Admin', is_active=True).distinct().count()
        if not target.is_active:
            target.is_active = True
            target.save()
            messages.success(request, f'{target.username} has been unlocked.')
        else:
            if target.groups.filter(name='Admin').exists() and admin_count <= 1:
                messages.error(request, 'Cannot lock the last admin.')
                return redirect('journal:user_management')
            target.is_active = False
            target.save()
            messages.success(request, f'{target.username} has been locked.')
        return redirect('journal:user_management')
